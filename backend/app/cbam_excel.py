import os
from datetime import date
from typing import List, Tuple
import openpyxl

from .models import Period, Product

def fill_cbam_template(template_path: str, period: Period, products: List[Product], out_path: str) -> str:
    wb = openpyxl.load_workbook(template_path)

    # --- A_InstData (installation + reporting period) ---
    ws = wb["A_InstData"]
    # Reporting period start/end
    ws["I9"].value = period.start_date
    ws["L9"].value = period.end_date

    # Installation fields
    ws["I19"].value = period.installation_name
    ws["I20"].value = period.installation_name_en
    ws["I21"].value = period.street_number
    ws["I22"].value = period.economic_activity
    ws["I23"].value = period.post_code
    ws["I24"].value = period.po_box
    ws["I25"].value = period.city
    ws["I26"].value = period.country
    ws["I27"].value = period.unlocode
    ws["I28"].value = period.latitude
    ws["I29"].value = period.longitude

    # --- C_Emissions&Energy (quality statements + indirect total placeholder) ---
    wsC = wb["C_Emissions&Energy"]
    wsC["H40"].value = period.data_quality
    wsC["H41"].value = period.default_values_justification
    wsC["H42"].value = period.quality_assurance

    # Total indirect emissions at installation level (manual entry cell M26)
    total_indirect = 0.0
    total_direct = 0.0
    for p in products:
        total_direct += (p.production_t or 0.0) * (p.direct_see or 0.0)
        total_indirect += (p.production_t or 0.0) * (p.indirect_see or 0.0)
    wsC["M26"].value = round(total_indirect, 6)

    # --- Summary_Products (direct fill for product lines) ---
    wsS = wb["Summary_Products"]
    start_row = 10
    # Clear first 200 rows (safety; only value cells)
    for r in range(start_row, start_row + 200):
        for col in ["D","E","F","G","H","I","J"]:
            wsS[f"{col}{r}"].value = None

    for i, p in enumerate(products):
        r = start_row + i
        wsS[f"D{r}"].value = "ISOTEC General process"
        wsS[f"E{r}"].value = p.aggregated_category or ""
        wsS[f"F{r}"].value = p.cn_code
        wsS[f"G{r}"].value = p.cn_name or ""
        wsS[f"H{r}"].value = p.product_name
        wsS[f"I{r}"].value = p.direct_see
        wsS[f"J{r}"].value = p.indirect_see
        # Total column K is formula; keep as is.

    wb.save(out_path)
    return out_path
