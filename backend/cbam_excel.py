"""
Excel report generator for CBAM reports.

This module contains helper functions to populate a pre‑defined CBAM
communication template with the data provided by the client. For the MVP
implementation the focus is on the "Summary_Products" sheet, where a
per‑product breakdown of the reporting period is written. In future
iterations additional sheets (e.g. A_InstData, C_Emissions&Energy) can be
populated from the report data model.

The template file ``cbam_template.xlsx`` must reside in
``backend/templates`` and will be copied into the report output directory.
"""

import os
from openpyxl import load_workbook
from typing import Any

from main import ReportRequest, Product  # type: ignore  # circular but safe during runtime


def generate_excel_report(report: ReportRequest, output_path: str) -> None:
    """Populate the CBAM template with report data and save it.

    Currently only the ``Summary_Products`` sheet is populated. Each
    product in the input list is written into successive rows starting at
    row 6. The columns map as follows:

    * Column A: CN code
    * Column B: Product name
    * Column C: Production in tonnes
    * Column D: Direct SEE (tCO2e/t)
    * Column E: Indirect SEE (tCO2e/t)
    * Column F: Total SEE (tCO2e/t)

    Args:
        report: Parsed request data containing the quarter and products.
        output_path: Where the generated Excel file should be stored.

    Raises:
        FileNotFoundError: If the template file cannot be located.
    """
    template_dir = os.path.join(os.path.dirname(__file__), "templates")
    template_path = os.path.join(template_dir, "cbam_template.xlsx")
    if not os.path.isfile(template_path):
        raise FileNotFoundError(
            "CBAM template not found. Expected at backend/templates/cbam_template.xlsx"
        )

    # Load template workbook
    wb = load_workbook(template_path)

    # Access the sheet where product summaries are listed
    if "Summary_Products" not in wb.sheetnames:
        # Fall back to the first sheet if the expected sheet is absent
        ws = wb[wb.sheetnames[0]]
    else:
        ws = wb["Summary_Products"]

    # Starting row for product entries; adjust if template layout changes
    start_row = 6

    for idx, product in enumerate(report.products):
        row = start_row + idx
        ws[f"A{row}"] = product.cn_code
        ws[f"B{row}"] = product.name
        ws[f"C{row}"] = product.production_ton
        ws[f"D{row}"] = product.direct_see
        ws[f"E{row}"] = product.indirect_see
        ws[f"F{row}"] = product.total_see

    # Optionally, write the reporting period in a known cell if the template supports it
    # For example, cell C2 could contain the quarter label (to be adjusted per template)
    try:
        ws["C2"] = report.quarter
    except Exception:
        # Silently ignore if the cell doesn't exist
        pass

    # Save a new file (not overwrite the template)
    wb.save(output_path)